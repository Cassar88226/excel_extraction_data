from PyQt5 import QtWidgets, uic, QtCore, QtGui
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal
import sys, os
import pandas as pd
import numpy as np
from datetime import datetime
import dateutil.relativedelta
import xlrd
from xlrd import open_workbook
from pandas import ExcelWriter
from openpyxl import load_workbook

# sheet_name = "Extract"

def display_error_message(content):
    msg = QMessageBox()
    msg.setIcon(QMessageBox.Critical)
    msg.setText("Error")
    msg.setInformativeText(content)
    msg.setWindowTitle("Error")
    msg.exec_()

class MainThread(QThread):
    change_value = pyqtSignal(int, str)
    def __init__(self):
        QThread.__init__(self)
        self.dst_filename = ""
        self.src_filenames = []

    # Create a counter thread
    def run(self):
        writer = ExcelWriter(self.dst_filename, engine='openpyxl')
        # create empty dataframe
        total_df = pd.DataFrame(columns=['Title', 'ISWC', 'SOCIETY', 'IP Name', 'IPN#', 'Role', 'P-Society', 'P-Share', 'M-Society', 'M-Share'])
        total_creator_df = pd.DataFrame(columns=['IP Name', 'IPN#', 'Role', 'P-Society', 'P-Share', 'M-Society', 'M-Share'])
        
        # if os.path.isfile(self.dst_filename):
        #     org_book = load_workbook(self.dst_filename)
        #     writer.book = org_book
        #     writer.sheets = dict((ws.title, ws) for ws in org_book.worksheets)
        self.change_value.emit(0, 'Creating New Dataset')
        try:
            for file in self.src_filenames:
                path, filename = os.path.split(file)
                self.change_value.emit(0, 'Processing file : ' + filename)
                wb = open_workbook(file)
                book = load_workbook(file)
                all_sheets = wb.sheet_names()


                # 1. Get New Dataset

                title_cnt = 0
                all_start_rows = []
                all_titles = []

                # get start rows of new title for each sheet
                for sheet_name in all_sheets:
                    row = 0
                    start_rows = []
                    titles = []
                    Extract_sheet = wb.sheet_by_name(sheet_name)
                    while row  < Extract_sheet.nrows:
                        # get first cell of each row
                        first_cell_of_row = Extract_sheet.cell_value(row, 0)
                        # if first cell of row is not empty, it means new Title starts from its row
                        title = ''
                        if first_cell_of_row:
                            start_rows.append(row)
                            title = str(Extract_sheet.cell_value(row, 3))
                            titles.append(title)
                            title_cnt += 1
                        row += 1
                    all_start_rows.append(start_rows)
                    all_titles.append(titles)
                title_index = 0
                for index, start_rows in enumerate(all_start_rows):
                    # sheet_df = pd.DataFrame(columns=['Title', 'ISWC', 'SOCIETY', 'IP Name', 'IPN#', 'Role', 'P-Society', 'P-Share', 'M-Society', 'M-Share'])
                    # sheet_creator_df = pd.DataFrame(columns=['IP Name', 'IPN#', 'Role', 'P-Society', 'P-Share', 'M-Society', 'M-Share'])
                    for index1, start_row in enumerate(start_rows):
                        Extract_sheet = wb.sheet_by_name(all_sheets[index])
                        end_row = 0
                        if index1 == len(start_rows) - 1:
                            end_row = Extract_sheet.nrows
                        else:
                            end_row = start_rows[index1 + 1]

                        # get ISWC, Society, Last Update
                        iswc = ''
                        society = ''
                        last_update = ''
                        for row1 in range(start_row + 1, end_row):
                            valid_row = False
                            for col in range(0, Extract_sheet.ncols):
                                cell_value = str(Extract_sheet.cell_value(row1, col))
                                if cell_value.strip():
                                    valid_row = True
                                if 'ISWC' in cell_value:
                                    iswc = Extract_sheet.cell_value(row1, col + 1)
                                if 'Submitting Society' in cell_value:
                                    society = str(Extract_sheet.cell_value(row1, col + 1))
                                if 'Society Work Code' in cell_value:
                                    society_code = str(Extract_sheet.cell_value(row1, col + 1))
                                if 'Last Update' in cell_value:
                                    last_update = Extract_sheet.cell_value(row1, col + 1)
                            # check if row is empty
                            if not valid_row:
                                row1 += 1
                                break
                        # find Creater dataframe
                        creater_st_row = row1 + 1
                        creater_en_row = row1 + 1
                        for row1 in range(creater_st_row, end_row):
                            next_paragraph_cell = Extract_sheet.cell_value(row1, 2)
                            if 'Creator(s)' in next_paragraph_cell:
                                creater_st_row = row1 + 1
                            if 'Publisher(s)' in next_paragraph_cell:
                                creater_en_row = row1 - 2
                                break
                        creator_df = pd.DataFrame(columns=['IP Name', 'IPN#', 'Role', 'P-Society', 'P-Share', 'M-Society', 'M-Share'])
                        for row1 in range(creater_st_row + 1, creater_en_row+1):
                            row_list = Extract_sheet.row_values(row1, start_colx=2, end_colx=None)
                            creator_df.loc[len(creator_df)] = row_list

                        # creator_df = pd.read_excel(excel_file_path, index_col = None, skiprows= creater_st_row, 
                        #     nrows= creater_en_row - creater_st_row, sheet_name=sheet_name, usecols=range(2,Extract_sheet.ncols),
                        #     converters=converters)

                        # assign Title, ISWC, SOCIETY in creator dataframe
                        creator_df = creator_df.assign(SOCIETY=society)[['SOCIETY'] + creator_df.columns.tolist()]
                        creator_df = creator_df.assign(ISWC=iswc)[['ISWC'] + creator_df.columns.tolist()]
                        creator_df = creator_df.assign(Title=all_titles[index][index1])[['Title'] + creator_df.columns.tolist()]
                        creator_df = creator_df.assign(LAST_UPDATE=last_update)

                        # merge Dataframes
                        total_df = pd.concat([total_df, creator_df])
                        # merge creator dataframe
                        creator_df = creator_df.assign(Title_No=title_cnt)[['Title_No'] + creator_df.columns.tolist()]
                        creator_df = creator_df.assign(Society_Code=society_code)[['Society_Code'] + creator_df.columns.tolist()]
                        total_creator_df = pd.concat([total_creator_df, creator_df])


                        # find Publisher dataframe
                        publisher_st_row = row1 + 1
                        publisher_en_row = row1 + 1
                        for row1 in range(publisher_st_row, end_row):
                            next_paragraph_cell = Extract_sheet.cell_value(row1, 2)
                            if 'Publisher(s)' in next_paragraph_cell:
                                publisher_st_row = row1 + 1
                            if 'Performer(s)' in next_paragraph_cell:
                                publisher_en_row = row1 - 2
                                break
                        publisher_df = pd.DataFrame(columns=['IP Name', 'IPN#', 'Role', 'P-Society', 'P-Share', 'M-Society', 'M-Share'])
                        for row1 in range(publisher_st_row + 1, publisher_en_row+1):
                            row_list = Extract_sheet.row_values(row1, start_colx=2, end_colx=None)
                            publisher_df.loc[len(publisher_df)] = row_list

                        # publisher_df = pd.read_excel(excel_file_path, index_col = None, skiprows= publisher_st_row, 
                        #     nrows= publisher_en_row - publisher_st_row, sheet_name=sheet_name, usecols=range(2,Extract_sheet.ncols), 
                        #     converters=converters)
                        # assign Title, ISWC, SOCIETY in publisher dataframe
                        publisher_df = publisher_df.assign(SOCIETY=society)[['SOCIETY'] + publisher_df.columns.tolist()]
                        publisher_df = publisher_df.assign(ISWC=iswc)[['ISWC'] + publisher_df.columns.tolist()]
                        publisher_df = publisher_df.assign(Title=all_titles[index][index1])[['Title'] + publisher_df.columns.tolist()]
                        publisher_df = publisher_df.assign(LAST_UPDATE =last_update)

                        # merge Dataframes
                        total_df = pd.concat([total_df, publisher_df])
                        title_index += 1
                        self.change_value.emit(int(title_index * 100 / title_cnt), 'Processing file : ' + filename)

            total_df.to_excel(writer,'New Dataset Python', index=False)

            # 2. Stage One
            self.change_value.emit(100, 'Processing of Stage 1')
            creator_group_df = total_creator_df.groupby('IP Name')['IPN#'].apply(set).reset_index(name='IPN list')
            stage_one_df = pd.DataFrame(columns=['IP Name', 'IPN#'])
            for i, row in creator_group_df.iterrows():
                IPN_list = list(row['IPN list'])
                if len(IPN_list) > 1 or (len(IPN_list) == 1 and np.nan in IPN_list):
                    for entry in IPN_list:
                        stage_one_df = stage_one_df.append({'IP Name':row['IP Name'], 'IPN#':entry}, ignore_index=True)
            stage_one_df = stage_one_df.replace(np.nan, 'Blank', regex=True)
            stage_one_df.to_excel(writer, 'Stage One', index=False)


            # 3. Stage Two
            self.change_value.emit(100, 'Processing of Stage 2')
            title_group_df = total_creator_df.groupby(['Title_No', 'Title', 'ISWC']).size().reset_index(name='counts')
            title_group_df = title_group_df.groupby(['Title', 'ISWC'])['counts'].apply(set).reset_index(name='count_list')
            # title_group_df['count_list'] = title_group_df['count_list'].astype(str)
            stage_two_df = pd.DataFrame(columns=['Title', 'No. of Creators', 'ISWC'])
            for i, row in title_group_df.iterrows():
                creator_count_list = list(row['count_list'])
                if len(creator_count_list) > 1:
                    for entry in creator_count_list:
                        stage_two_df = stage_two_df.append({'Title':row['Title'], 'No. of Creators':entry, 'ISWC':row['ISWC']}, ignore_index=True)
            stage_two_df.to_excel(writer, 'Stage Two', index=False)

            # 4. Stage Three
            self.change_value.emit(100, 'Processing of Stage 3')
            stage_three_df = total_creator_df.loc[(total_creator_df['ISWC'] == 'No preferred') | (total_creator_df['ISWC'] == '')]
            stage_three_df = stage_three_df.groupby(['Title', 'Society_Code']).size().reset_index(name='counts').drop(['counts'], axis=1)
            stage_three_df.to_excel(writer, 'Stage Three', index=False)

            # 5. Stage Four
            self.change_value.emit(100, 'Processing of Stage 4')
            end_time = datetime.now()
            end_date = end_time.strftime('%Y/%m/%d')
            start_time = end_time - dateutil.relativedelta.relativedelta(months=2)
            start_date = start_time.strftime('%Y/%m/%d')
            mask = ((total_creator_df['LAST_UPDATE'] > start_date) & (total_creator_df['LAST_UPDATE'] <= end_date)) |(total_creator_df['LAST_UPDATE'] == '')
            stage_four_df = total_creator_df.loc[mask].groupby(['Title', 'LAST_UPDATE', 'ISWC']).size().reset_index(name='counts').drop(['counts'], axis=1)
            stage_four_df.to_excel(writer, 'Stage Four', index=False)

            # 6. Stage Five
            self.change_value.emit(100, 'Processing of Stage 5')
            # group by title and creator
            shared_df = total_creator_df.loc[total_creator_df['P-Share'] != '*']
            shared_df = shared_df.copy()
            shared_df['P-Share'] = shared_df['P-Share'].astype('float')
            shared_df = shared_df.groupby(['Title_No', 'Title', 'IP Name'])['P-Share'].sum().reset_index(name='%')#.apply(list)
            stage_five_df = shared_df.drop_duplicates(subset=['Title', 'IP Name', '%'], keep=False).drop(['Title_No'], axis=1)
            stage_five_df = stage_five_df.rename(columns={'IP Name':'Creator'})
            stage_five_df.to_excel(writer, 'Stage Five', index=False)
            
            # 7. Save all sheets
            writer.save()
            self.change_value.emit(100, 'Completed')
        except:
            self.change_value.emit(0, 'Error occurred in input files')
            # display_error_message("Input file format error")
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        uic.loadUi("main.ui", self)
        self.centralwidget.setContentsMargins(10, 10, 10, 10);
        # self.setFixedSize(722, 532)
        self.selected_rows = []
        self.src_filenames = []
        self.dst_filename = ""
        self.prgressbar_Run.setValue(0)
    
    @QtCore.pyqtSlot()
    def selecteditems(self):
        cur_rows = [x.row() for x in self.list_files.selectedIndexes()]
        self.selected_rows = cur_rows

    @QtCore.pyqtSlot()
    def addfiles(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileNames, _ = QFileDialog.getOpenFileNames(self,"QFileDialog.getOpenFileName()", "","Excel Files (*.xlsx)", options=options)
        if fileNames:
            for file in fileNames:
                self.list_files.addItem(file)

    @QtCore.pyqtSlot()
    def deletefiles(self):
        for row in self.selected_rows:
            self.list_files.takeItem(row)
            

    @QtCore.pyqtSlot()
    def browse(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, filter = QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","Excel Files (*.xlsx)", options=options)
        if fileName:
            self.dst_filename = fileName
            if not fileName.endswith('.xlsx'):
                self.dst_filename = fileName + ".xlsx"
            self.edit_output_file.setText(self.dst_filename)
    @QtCore.pyqtSlot()
    def run(self):
        if self.edit_output_file.text():
            self.src_filenames = [self.list_files.item(i).text() for i in range(self.list_files.count())]
            if self.src_filenames:
                self.main_thread = MainThread()
                self.main_thread.dst_filename = self.dst_filename
                self.main_thread.src_filenames = self.src_filenames
                self.main_thread.change_value.connect(self.setProgressVal)
                self.main_thread.start()
            else:
                display_error_message("Please put input files")
        else:
            display_error_message("Please put output file")

    def setProgressVal(self, val, text):
        self.prgressbar_Run.setValue(val)
        self.lbl_progress.setText(text)


if __name__ == '__main__':
    app = QtWidgets.QApplication([])
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())
